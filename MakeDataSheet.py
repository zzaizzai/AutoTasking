import pandas as pd
import os
import openpyxl
import Service

def make_data_sheet(target:str):

    def CellWidth(work_book):
        print('fixing cell width')

        wb = work_book
        ws = wb.worksheets[0]

        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20

        return wb

    Data_sheet_path = Service.data_dir(target) + fr'\{target} Data Sheet.xlsx'
    Data_path = Service.data_dir(target) + fr'\{target} Data.xlsx'

    if os.path.isfile(Data_path):
        print('data file exist')

    else:
        print("no data file ")
        return

    df = pd.read_excel(Data_path, index_col=False)
    target_list = [x for x in df.columns.to_list() if x[3:].isdigit()]
    print(target_list)


    title_list = ["method", "condition", "type", "unit"] + target_list

    wb = openpyxl.Workbook()
    ws = wb.active 
    row_start = 4
    col_start = 2
    # ws.cell(row_start, col_start).value = 

    # make titles
    col_title = col_start
    for title in title_list:
        ws.cell(row_start, col_title).value = title
        if col_title >= col_start + 4:
            col_title += 2
        else:
            col_title += 1


    condition_now = "123"
    row_index = row_start + 1
    for method in ["ムーニー_ロータ_自動集積", "レオメータ", "初期物性", "硬度_自動集積 ","熱老化","耐油引張り", "⊿Ｖ", 'ΔV', "圧縮永久歪", '脆化',"押出し", "オゾン"]:
        method_count = 0
        # print(method)
        for i, value in enumerate(df["method"].str.contains(method)):
            col_index = col_start
            # print(i, value)
            
            if value:
                row = df.iloc[i, :]
                # print(row["method"], row["condition"], row["type"], row["unit"])
                for title in title_list:

                    if method_count > 0 and title == "method":
                        pass
                    elif row[title] == condition_now and title == "condition":
                        pass
                    else:
                        ws.cell(row_index, col_index).value = row[title]
                    
                    if title == "condition":
                        condition_now = row["condition"]


                    method_count += 1
                    if col_index >= col_start + 4:
                        col_index += 2
                    else:
                        col_index += 1
                row_index += 1

    # print(df["method"].str.contains("レオメータ"))

    # width of wrokbook
    wb = CellWidth(wb)

    wb.save(Data_sheet_path)
    


def Doit(target:str):
    try:
        make_data_sheet(target)
    except Exception as e:
        print(e)

if __name__ == "__main__":
    print("make data sheet")
    target = input("input your target: ")
    Doit(target)