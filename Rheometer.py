import os
import glob
import win32com.client as win32
import pandas as pd
import openpyxl
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties, Font
from openpyxl.chart.layout import Layout, ManualLayout
import Service


class Rheometer:

    test_mode = False

    def TestMode(self, mode: bool):
        self.TestMode = mode

    def __init__(self, target):
        self.exp_name = 'レオメータ'

        self.file_dir = Service.data_dir(target)
        self.target = target
        self.file_path_xls = self.file_dir + \
            rf'\{self.exp_name}*{target}*.xlsx'
        self.file_xls = ''
        self.file_now = ''
        self.temperature = ''
        self.standard_cell = ''

    def StartProcess(self):
        print('find file rheometer...')
        print(self.file_path_xls)

        file_list = glob.glob(self.file_path_xls)
        file_list = sorted(file_list, key=len)
        print(file_list)

        if len(file_list) > 0:
            print(f'found {len(file_list)} {self.exp_name} file(s) ')

            for file in file_list:
                self.file_now = file
                self.MakeXlsmFile()
        else:
            print(f'No {self.exp_name}')
            return

    def MakeXlsmFile(self):

        self.CreateGraph()
        self.ReadFile()

    def CreateGraph(self):
        print('creating graph')
        wb = openpyxl.load_workbook(self.file_now)
        ws = wb.worksheets[0]

        standard_cell = ws.cell(row=1, column=1)

        # Find 'Time' value of cell

        is_ok = False
        for row in ws.rows:
            for cell in row:
                if cell.value == 'Time(NO.1)':
                    standard_cell = cell
                    self.standard_cell = cell
                    is_ok = True
                    break
            if is_ok:
                break

        self.number_of_target = int(
            ws.cell(row=standard_cell.row - 2,
                    column=standard_cell.column).value)
        print(f'number of target: {self.number_of_target}')

        print(f'standard postion cell: {standard_cell}')

        # create a graph
        chart = openpyxl.chart.ScatterChart('marker')
        # chart = openpyxl.chart.LineChart()

        # data titles

        temperature_cell = ws.cell(row=2, column=6)

        temperature_title = str(int(float(temperature_cell.value[:-1]))) + '℃'
        self.temperature = temperature_title
        # print(temperature_title)

        chart.title = f'加硫曲線 {temperature_title}'
        font = Font(typeface='Calibri')
        size = 1600
        char_prop = CharacterProperties(latin=font, sz=size, b=True)
        para_prop = ParagraphProperties(defRPr=char_prop)
        chart.title.tx.rich.p[0].pPr = para_prop

        chart.x_axis.title = 'Time (min)'
        chart.y_axis.title = 'Torque (kgf・cm)'

        font = Font(typeface='Calibri')
        size = 1200  # 14 point size
        char_prop = CharacterProperties(latin=font, sz=size, b=False)
        para_prop = ParagraphProperties(defRPr=char_prop)
        chart.x_axis.title.tx.rich.p[0].pPr = para_prop
        chart.y_axis.title.tx.rich.p[0].pPr = para_prop

        colors = {
            'blue': '0000ff',
            'red': 'ff0000',
            'green': '00cc00',
            'orange': 'ff8000',
            'violet': 'ff00ff',
            'pink': 'ff00d5',
        }
        colors = [
            colors['red'], colors['blue'], colors['green'], colors['orange'],
            colors['violet'], colors['pink'], colors['pink'], colors['pink'],
            colors['pink'], colors['pink']
        ]
        # add series
        if self.test_mode:
            print('adding serials')

        for i in range(self.number_of_target):
            serial_name = ws.cell(row=self.standard_cell.row,
                                  column=2 + 4 * i).value
            if serial_name != None:

                if self.test_mode:
                    print(serial_name)

                data = openpyxl.chart.Reference(ws,
                                                min_col=2 + i * 4,
                                                min_row=standard_cell.row,
                                                max_row=1000)
                times = openpyxl.chart.Reference(ws,
                                                 min_col=1,
                                                 min_row=standard_cell.row + 1,
                                                 max_row=1000)

                series = openpyxl.chart.Series(data,
                                               times,
                                               title_from_data=True)
                # series.graphicalProperties.line.noFill = True

                series.graphicalProperties.line.solidFill = colors[i]

                # series.marker.symbol = "circle"
                # series.marker.size = 0
                # series.spPr.ln.solidFill = "000000"
                if i == 0:
                    series.spPr.ln.prstDash = "sysDash"
                else:
                    pass

                series.smooth = True

                chart.series.append(series)

        if self.test_mode:
            print(f'number of series: {len(chart.series)}')

        # y scale
        # chart.y_axis.scaling.min = 0
        # chart.y_axis.scaling.max = 35

        # x scale
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 45

        chart.legend.position = 'r'
        # chart.layout = Layout(
        #     ManualLayout(
        #     x=0.5, y=0.5,
        #     h=0.5, w=0.5,
        #     xMode="edge",
        #     yMode="edge",
        #     )
        # )
        # chart size
        chart.height = 12
        chart.width = 12

        def target_number(number: int):
            target = self.target
            alphabet = target[0:3]
            num = int(target[3:])
            alphabet_num = alphabet + str('%03d' % (num + number))
            return alphabet_num

        for i in range(self.number_of_target):
            ws.cell(row=standard_cell.row,
                    column=2 + i * 4).value = target_number(i)

        ws.add_chart(chart, 'B8')
        wb.save(self.file_now)

        print('graph save done!!')

    def ReadFile(self):
        df = pd.read_excel(self.file_now, header=None)

        num_target = 0
        row_init = 0

        for i, value in enumerate(df[0]):
            if value == '特性値：':
                num_target = df[0][i - 1]
                row_init = i + 4

        if self.test_mode:
            print(f'number of target: {num_target}')
            print(f'samples start row: {row_init}')

        df_input = df.loc[[row_init]]
        for i in range(1, num_target):
            df_input = df_input.append(df.loc[[row_init + 2 * i]])

        df_input = df_input.transpose()

        df_input = df_input.loc[[2, 3, 4, 5, 6, 7, 8]]

        print('changing target numbers')

        target_title = []
        for i in range(len(df_input.columns)):
            target_title.append(Service.target_number(i, self.target))
        print(target_title)
        df_input.columns = target_title

        unit = ['kgf・cm', 'kgf・cm', 'min', 'min', 'min', 'min', 'min']
        type_list = ['MH', 'ML', 'ts1', 't10', 't50', 't90', 'CR']

        condition_environment = ""
        if Service.file_name_without_target_and_expname(
                self.file_now, self.target, self.exp_name) == "none":
            pass
        else:
            condition_environment = Service.file_name_without_target_and_expname(
                self.file_now, self.target, self.exp_name)

        condition = [f'{self.temperature} {condition_environment}'
                     ] * len(df_input)
        # something is needed in condition.....
        method_list = [
            Service.file_name_without_target(self.file_now, self.target)
        ] * len(df_input)

        df_input.insert(0, 'unit', unit)
        df_input.insert(0, 'type', type_list)
        df_input.insert(0, 'condition', condition)
        df_input.insert(0, 'method', method_list)

        self.WriteData(df_input)

    def WriteData(self, df_input):
        print('writing data....')

        file_data = self.file_dir + fr'\{self.target} Data.xlsx'

        is_file = os.path.isfile(file_data)

        if is_file:
            pass
        else:
            print('no data file')
            # or you can make a data file
            return

        Service.save_to_data_excel(file_data, df_input, self.exp_name)


def DoIt(target: str, test_mode=False):
    reo = Rheometer(target)
    reo.TestMode(test_mode)
    try:
        reo.StartProcess()
    except Exception as e:
        print(e)


if __name__ == '__main__':
    target = input('target number (ex: ABC001): ')
    DoIt(target, test_mode=True)
