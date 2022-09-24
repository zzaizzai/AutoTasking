import os
import glob
import win32com.client as win32
import pandas as pd
import openpyxl
from openpyxl.drawing.text import ParagraphProperties, CharacterProperties, Font

class Rheometer:

    def __init__(self, target):
        self.exp_name = 'レオメータ'

        self.DesktopPath = os.path.expanduser('~/Desktop')

        self.file_dir = self.DesktopPath + rf'\{target} Data'
        self.target = target
        self.file_path_xls = self.file_dir + \
            rf'\{self.exp_name}*{target}*.xlsx'
        self.file_xls = ''
        self.file_xlsx = ''

    def FindFile(self):
        print('find file rheometer...')
        print(self.file_path_xls)

        file_list = glob.glob(self.file_path_xls)
        file_list = sorted(file_list, key=len)
        print(file_list)

        if len(file_list) > 0:
            print(f'found {len(file_list)} {self.exp_name} file(s) ')

            for file in file_list:
                self.file_xlsx = file
                self.MakeXlsmFile()
        else:
            print(f'No {self.exp_name}')
            return

    def MakeXlsmFile(self):
        # print('make xlsm file...')

        # self.file_xlsx = self.file_xls + 'x'
        # is_file = os.path.isfile(self.file_xlsx)

        # if is_file:
        #     print('xlsx file exist')
        #     os.remove(self.file_xlsx)
        # else:
        #     pass

        # print('make new xlsx file...')
        # excel = win32.gencache.EnsureDispatch('Excel.Application')
        # wb = excel.Workbooks.Open(self.file_xls)
        # wb.SaveAs(self.file_xlsx, FileFormat=51)
        # wb.Close()
        # excel.Application.Quit()

        # print('make new xlsx file done')

        self.CreateGraph()
        self.ReadFile()

    def CreateGraph(self):
        print('creating graph')
        wb = openpyxl.load_workbook(self.file_xlsx)
        ws = wb.worksheets[0]

        standard_cell = ws.cell(row=1, column=1)

        # Find 'Time' value of cell

        is_ok = False
        for row in ws.rows:
            for cell in row:
                if cell.value == 'Time(NO.1)':
                    standard_cell = cell
                    is_ok = True
                    break
            if is_ok:
                break

        self.number_of_target = int(
            ws.cell(row=standard_cell.row - 2, column=standard_cell.column).value)
        print(f'number of target: {self.number_of_target}')

        print(f'standard postion cell: {standard_cell}')
        # 1/4 data
        # print('1/4 data....')
        # from_row = standard_cell.row + 20
        # for i in range(200):
        #     print(f'deleting{i}')
        #     ws.delete_rows( from_row + i*3 + 3)
        #     ws.delete_rows( from_row + i*3 + 2)
        #     ws.delete_rows( from_row + i*3 + 1)
        #     ws.delete_rows( from_row + i*3)

        # create a graph
        chart = openpyxl.chart.ScatterChart('marker')
        # chart = openpyxl.chart.LineChart()

        # data titles
        chart.title = 'Rheometer'
        font = Font(typeface='Calibri')
        size = 1600
        char_prop = CharacterProperties(latin=font, sz=size, b=True) 
        para_prop = ParagraphProperties(defRPr=char_prop)
        chart.title.tx.rich.p[0].pPr = para_prop


        chart.x_axis.title = 'Time (min)'
        chart.y_axis.title = 'torque'

        
        font = Font(typeface='Calibri')
        size = 1200 # 14 point size 
        char_prop = CharacterProperties(latin=font, sz=size, b=False) 
        para_prop = ParagraphProperties(defRPr=char_prop)
        chart.x_axis.title.tx.rich.p[0].pPr = para_prop
        chart.y_axis.title.tx.rich.p[0].pPr = para_prop


        colors  = ['ff0000','ffaa00','00ff00' ,'0055ff','8000ff','00ff55','ff00d5']
        # add series
        print('adding serials')
        for i in range(self.number_of_target):
            serial_name = ws.cell(row=21, column=2 + 4*i).value
            if serial_name != None:
                print(serial_name)
                data = openpyxl.chart.Reference(
                    ws, min_col=2 + i*4, min_row=standard_cell.row, max_row=1000)
                times = openpyxl.chart.Reference(
                    ws, min_col=1, min_row=standard_cell.row + 1, max_row=1000)

                series = openpyxl.chart.Series(
                    data, times, title_from_data=True)
                # series.graphicalProperties.line.noFill = True
                
                series.graphicalProperties.line.solidFill = colors[i]


                # series.marker.symbol = "circle"
                # series.marker.size = 0
                # series.spPr.ln.solidFill = "000000"
                # series.smooth = True

                chart.series.append(series)

        print(f'number of series: {len(chart.series)}')

        # # data scail
        # chart.y_axis.scaling.min = 0
        # chart.y_axis.scaling.max = 35

        # chart.x_axis.scaling.min = 0
        # chart.x_axis.scaling.max = 30

        ws.add_chart(chart, 'B8')
        wb.save(self.file_xlsx)

        print('graph save done!!')

    def ReadFile(self):
        df = pd.read_excel(self.file_xlsx, header=None)

        num_target = 0
        row_init = 0
        print(df)
        for i, value in enumerate(df[0]):
            if value == '特性値：':
                num_target = df[0][i-1]
                row_init = i + 4

        print(f'number of target: {num_target}')
        print(f'samples start row: {row_init}')

        df_input = df.loc[[row_init]]
        for i in range(1, num_target):
            df_input = df_input.append(df.loc[[row_init + 2*i]])

        df_input = df_input.transpose()
        print(df_input)
        df_input = df_input.loc[[2, 3, 5, 6, 7, 8]]

        print(df_input)

        file_name = os.path.splitext(os.path.basename(self.file_xlsx))[0]
        unit = ['kgf・cm', 'kgf・cm', 'min', 'min', 'min', 'min']
        method = ['MH', 'ML', 't10', 't50', 't90', 'CR']
        condition = ['none','none','none','none','none','none',]
        name = [file_name, file_name, file_name,
                file_name, file_name, file_name, ]
        df_input.insert(0, 3, unit)
        df_input.insert(0, 2, method)
        df_input.insert(0, 1, condition)
        df_input.insert(0, 0, name)

        print(df_input)

        # reset title and index
        df_input.reset_index(inplace=True, drop=True)
        df_input = df_input.T.reset_index(drop=True).T

        print(df_input)

        self.WriteData(df_input)

    def WriteData(self, df_input):
        print('writing data....')

        file_data = self.file_dir + fr'\{self.target} Data.xlsx'

        is_file = os.path.isfile(file_data)

        if is_file:
            pass
        else:
            print('no data file')
            # or you can make a data file
            return

        df = pd.read_excel(file_data, index_col=0)
        print(df)

        df_merge = pd.concat([df, df_input])
        print(df_merge)

        df_merge.reset_index(inplace= True, drop= True)

        df_merge.to_excel(file_data, index=True, header=True, startcol=0)
        print(f'saved data file in {file_data}')


def Rheomeo(target: str):
    reo = Rheometer(target)
    reo.FindFile()


if __name__ == '__main__':
    target = input('target number (ex: ABC001): ')
    Rheomeo(target)
