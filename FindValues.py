import openpyxl

target = 'CBA001'
filePath = fr'C:\Users\junsa\Desktop\CBA001\Oil\{target}.xlsm'



def FindCell(rowValue, colValue, path: str):
    wb = openpyxl.load_workbook(path)
    sheet = wb.worksheets[0]
    for row in sheet.rows:
        for cell in row:
            if cell.value == rowValue:
                x = cell.row
            if cell.value == colValue:
                y = cell.column
    print(f'row: {x}, col: {y}, value: {sheet.cell(row= x, column= y).value}')
    wb.close()
    return sheet.cell(row= x, column= y)


print(FindCell(0.13, 'N4', filePath).value)